# Build an upgraded version of the project that was defended last night. When you have extracted the email, 
# phone number and links from the website; store the data on excel, forward the file to the email address of 
# the user.When you run this application it should say extraction successful. Also the application should ask 
# you to enter the email address and then send.

import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

def web_scrap():
    url = input("Enter the website URL: ")
    filename = input("Enter the filename to save the emails: ")

    if not url.startswith("http://") and not url.startswith("https://"):
        print("Please enter a valid website URL.")
        return web_scrap()

    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    phone_numbers = re.findall(r'\+?\d{10}(?:\s+\d{3}-\d{3}-\d{4})?\b', soup.get_text())
    emails = re.findall(r"[A-Za-z0-9%_+-.]+"
                        r"@[A-Za-z0-9.-]+"
                        r"\.[A-Za-z:;*$#]{2,6}", soup.get_text())
    links = [link.get('href') for link in soup.find_all('a', href=True)]

    print("Phone Numbers:")
    print(phone_numbers)
    print("Emails:")
    print(emails)
    print("Links:")
    print(links)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for i, email in enumerate(emails, start=1):
        sheet.cell(row=i, column=1, value=email)
    for i, phone_number in enumerate(phone_numbers, start=1):
        sheet.cell(row=i, column=2, value=phone_number)
    for i, link in enumerate(links, start=1):
        sheet.cell(row=i, column=3, value=link)

    excel_file = os.path.join(sys.path[0], filename)
    workbook.save(excel_file)

    print(f"Data saved to '{excel_file}'.")
    print("Extraction successful")

    return excel_file

def send_email_with_attachment(sender, recipient, password, subject, attachment_path):
    # Create the email message using MIMEMultipart()
    email = MIMEMultipart()
    email["Subject"] = subject
    email["From"] = sender
    email["To"] = recipient

    # Open the document to be attached
    with open(attachment_path, "rb") as attachment_file:
        file_content = attachment_file.read()

    # Create a MIME attachment object
    attachment = MIMEBase("application", "octet-stream")
    attachment.set_payload(file_content)

    # Encode the attachment in base64
    encoders.encode_base64(attachment)
    attachment.add_header("Content-Disposition", "attachment", filename=os.path.basename(attachment_path))

    # Add the attachment to the email message
    email.attach(attachment)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender, password)
        smtp.sendmail(sender, recipient, email.as_string())

    print("Attachment sent!")

if __name__ == '_main_':
    # Scrape the website and save data to an Excel file
    excel_file = web_scrap()

    # Set email parameters
    sender_email = "emchadexglobal@gmail.com"
    recipient_email = input("Enter the recipient's email address: ")
    password = "ewhkaqtxojttbbub"
    subject = "Web Scrapping and attachment email in Python"

    # Send the email with the attachment
    send_email_with_attachment(sender_email, recipient_email, password, subject, excel_file)